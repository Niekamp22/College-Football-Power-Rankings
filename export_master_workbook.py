from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_RATINGS_PATH = Path("output/cfbd_power_ratings_2025.csv")
DEFAULT_BACKTEST_PATH = Path("output/backtests/weekly_backtest_2025_regular.csv")
DEFAULT_BACKTEST_GAMES_PATH = Path("output/backtests/game_backtest_2025_regular.csv")
DEFAULT_WIN_TOTALS_PATH = Path("output/projections/projected_win_totals_2026.csv")
DEFAULT_PROJECTED_GAMES_PATH = Path("output/projections/projected_games_2026.csv")
DEFAULT_SCHEDULE_COVERAGE_PATH = Path("output/projections/schedule_coverage_2026.csv")
DEFAULT_ODDS_COMPARISON_PATH = Path("output/odds/ncaaf_game_odds_comparison.csv")
DEFAULT_OUTPUT_PATH = Path("output/power_ratings_master.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export a master workbook for the power ratings project.")
    parser.add_argument("--ratings", type=Path, default=DEFAULT_RATINGS_PATH)
    parser.add_argument("--backtest", type=Path, default=DEFAULT_BACKTEST_PATH)
    parser.add_argument("--backtest-games", type=Path, default=DEFAULT_BACKTEST_GAMES_PATH)
    parser.add_argument("--win-totals", type=Path, default=DEFAULT_WIN_TOTALS_PATH)
    parser.add_argument("--projected-games", type=Path, default=DEFAULT_PROJECTED_GAMES_PATH)
    parser.add_argument("--schedule-coverage", type=Path, default=DEFAULT_SCHEDULE_COVERAGE_PATH)
    parser.add_argument("--odds-comparison", type=Path, default=DEFAULT_ODDS_COMPARISON_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    return parser.parse_args()


def load_csv(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def autosize(worksheet) -> None:
    for column in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column]
        worksheet.column_dimensions[column[0].column_letter].width = min(max(len(value) for value in values) + 2, 34)


def write_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title)
    if not rows:
        worksheet["A1"] = "No data available"
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    autosize(worksheet)


def add_matchup_sheet(workbook: Workbook, ratings: list[dict[str, Any]]) -> None:
    teams_sheet = workbook.create_sheet("Teams")
    teams_sheet.append(["team", "rating"])
    teams_sheet["A1"].font = Font(bold=True)
    teams_sheet["B1"].font = Font(bold=True)
    for row in ratings:
        teams_sheet.append([row["team"], row["rating"]])
    autosize(teams_sheet)
    teams_sheet.sheet_state = "hidden"

    sheet = workbook.create_sheet("MatchupTool")
    sheet["A1"] = "Matchup Tool"
    sheet["A1"].font = Font(bold=True)
    sheet["A3"] = "Team A"
    sheet["A4"] = "Team B"
    sheet["A5"] = "Site"
    sheet["A7"] = "Team A Rating"
    sheet["A8"] = "Team B Rating"
    sheet["A9"] = "Site Adjusted Spread"
    sheet["A10"] = "Team A Win Probability"
    sheet["A11"] = "Team B Win Probability"

    sheet["B3"] = ratings[0]["team"] if ratings else ""
    sheet["B4"] = ratings[1]["team"] if len(ratings) > 1 else ""
    sheet["B5"] = "Neutral"
    sheet["B7"] = '=IFERROR(INDEX(Teams!$B:$B,MATCH(B3,Teams!$A:$A,0)),"")'
    sheet["B8"] = '=IFERROR(INDEX(Teams!$B:$B,MATCH(B4,Teams!$A:$A,0)),"")'
    sheet["B9"] = '=IF(OR(B7="",B8=""),"",B7-B8+IF(B5="Team A Home",2.5,IF(B5="Team B Home",-2.5,0)))'
    sheet["B10"] = '=IF(B9="","",NORM.S.DIST(B9/16,TRUE))'
    sheet["B11"] = '=IF(B10="","",1-B10)'

    team_count = len(ratings) + 1
    team_validation = DataValidation(type="list", formula1=f"=Teams!$A$2:$A${team_count}", allow_blank=False)
    site_validation = DataValidation(type="list", formula1='"Neutral,Team A Home,Team B Home"', allow_blank=False)
    sheet.add_data_validation(team_validation)
    sheet.add_data_validation(site_validation)
    team_validation.add(sheet["B3"])
    team_validation.add(sheet["B4"])
    site_validation.add(sheet["B5"])
    autosize(sheet)


def main() -> None:
    args = parse_args()
    ratings = load_csv(args.ratings)
    weekly_backtest = load_csv(args.backtest)
    game_backtest = load_csv(args.backtest_games)
    win_totals = load_csv(args.win_totals)
    projected_games = load_csv(args.projected_games)
    schedule_coverage = load_csv(args.schedule_coverage)
    odds_comparison = load_csv(args.odds_comparison)

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(workbook, "Rankings", ratings)
    add_matchup_sheet(workbook, ratings)
    write_sheet(workbook, "ProjectedWins", win_totals)
    write_sheet(workbook, "WeeklyMatchups", projected_games)
    write_sheet(workbook, "ScheduleCoverage", schedule_coverage)
    write_sheet(workbook, "OddsEdges", odds_comparison)
    write_sheet(workbook, "BacktestWeekly", weekly_backtest)
    write_sheet(workbook, "BacktestGames", game_backtest)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Saved master workbook to {args.output}")


if __name__ == "__main__":
    main()
