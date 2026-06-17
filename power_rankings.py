from __future__ import annotations

import argparse
import csv
import json
import math
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_FEATURES_PATH = Path("data/cfbd/processed/2025/team_features.csv")
DEFAULT_LINES_PATH = Path("data/cfbd/raw/2025/lines.json")

DEFAULT_CALIBRATION_FEATURES = [
    "latest_team_postgame_elo",
    "avg_team_postgame_elo",
    "avg_team_pregame_elo",
    "offense_ppa",
    "defense_ppa_allowed",
    "offense_success_rate",
    "defense_success_rate_allowed",
    "wepa_total",
    "wepa_allowed_total",
    "recent_avg_cover_margin",
    "avg_cover_margin",
    "recent_win_pct",
    "recent_fbs_win_pct",
    "recent_fbs_avg_margin",
    "fbs_avg_margin",
    "avg_opponent_postgame_elo",
    "fbs_avg_opponent_pregame_elo",
    "talent",
]

RIDGE_ALPHA = 0.75
PRIOR_DECAY_GAMES = 2.5
MIN_PRIOR_WEIGHT = 0.0
PRIOR_SCALE = 14.0

LOWER_IS_BETTER = {
    "defense_ppa_allowed",
    "defense_success_rate_allowed",
    "wepa_allowed_total",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build market-calibrated college football power ratings.")
    parser.add_argument("--features", type=Path, default=DEFAULT_FEATURES_PATH, help="Path to the processed feature CSV.")
    parser.add_argument("--lines", type=Path, default=DEFAULT_LINES_PATH, help="Path to the CFBD lines JSON file.")
    parser.add_argument("--top", type=int, default=25, help="Number of teams to print.")
    parser.add_argument("--save", type=Path, help="Optional path to save the full rankings as CSV.")
    parser.add_argument("--excel", type=Path, help="Optional path to save an Excel workbook.")
    parser.add_argument("--team-a", help="Optional team name for a neutral-field matchup query.")
    parser.add_argument("--team-b", help="Optional team name for a neutral-field matchup query.")
    return parser.parse_args()


def parse_value(value: str) -> Any:
    if value == "":
        return None
    try:
        numeric = float(value)
    except ValueError:
        return value
    if numeric.is_integer():
        return int(numeric)
    return numeric


def load_team_features(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = [{key: parse_value(value) for key, value in row.items()} for row in reader]
    return [row for row in rows if row.get("classification") == "fbs"]


def load_lines(path: Path) -> list[dict[str, Any]]:
    return json.loads(path.read_text(encoding="utf-8"))


def z_scores(rows: list[dict[str, Any]], field: str, reverse: bool = False) -> dict[str, float]:
    values = [float(row[field]) for row in rows if row.get(field) is not None]
    if not values:
        return {str(row["team"]): 0.0 for row in rows}

    mean = sum(values) / len(values)
    variance = sum((value - mean) ** 2 for value in values) / len(values)
    std_dev = math.sqrt(variance)
    if math.isclose(std_dev, 0.0):
        return {str(row["team"]): 0.0 for row in rows}

    scores: dict[str, float] = {}
    for row in rows:
        team = str(row["team"])
        value = row.get(field)
        if value is None:
            scores[team] = 0.0
            continue
        z_value = (float(value) - mean) / std_dev
        scores[team] = -z_value if reverse else z_value
    return scores


def build_team_vectors(
    rows: list[dict[str, Any]],
    calibration_features: list[str] = DEFAULT_CALIBRATION_FEATURES,
) -> dict[str, np.ndarray]:
    feature_z_scores = {
        field: z_scores(rows, field, reverse=field in LOWER_IS_BETTER)
        for field in calibration_features
    }

    vectors: dict[str, np.ndarray] = {}
    for row in rows:
        team = str(row["team"])
        vectors[team] = np.array([feature_z_scores[field][team] for field in calibration_features], dtype=float)
    return vectors


def average_home_spread(game: dict[str, Any]) -> float | None:
    spreads = [provider.get("spread") for provider in game.get("lines", []) if provider.get("spread") is not None]
    if not spreads:
        return None
    return sum(float(spread) for spread in spreads) / len(spreads)


def fit_market_model(
    rows: list[dict[str, Any]],
    lines: list[dict[str, Any]],
    calibration_features: list[str] = DEFAULT_CALIBRATION_FEATURES,
    ridge_alpha: float = RIDGE_ALPHA,
) -> tuple[np.ndarray, float, list[dict[str, Any]], dict[str, np.ndarray], dict[str, Any]]:
    team_vectors = build_team_vectors(rows, calibration_features=calibration_features)
    team_lookup = {str(row["team"]): row for row in rows}

    x_rows: list[np.ndarray] = []
    y_values: list[float] = []
    samples: list[dict[str, Any]] = []

    for game in lines:
        if game.get("homeClassification") != "fbs" or game.get("awayClassification") != "fbs":
            continue
        home_team = game.get("homeTeam")
        away_team = game.get("awayTeam")
        if home_team not in team_vectors or away_team not in team_vectors:
            continue

        home_spread = average_home_spread(game)
        if home_spread is None:
            continue

        x_rows.append(team_vectors[home_team] - team_vectors[away_team])
        y_values.append(-home_spread)
        samples.append(
            {
                "week": game.get("week"),
                "home_team": home_team,
                "away_team": away_team,
                "market_home_line": round(float(home_spread), 3),
                "market_home_margin": round(float(-home_spread), 3),
                "home_team_record": f"{team_lookup[home_team].get('wins', 0)}-{team_lookup[home_team].get('losses', 0)}",
                "away_team_record": f"{team_lookup[away_team].get('wins', 0)}-{team_lookup[away_team].get('losses', 0)}",
            }
        )

    if not x_rows:
        raise SystemExit("No FBS line data was available to calibrate the market model.")

    x_matrix = np.vstack(x_rows)
    y_vector = np.array(y_values, dtype=float)
    design_matrix = np.column_stack([np.ones(len(x_matrix)), x_matrix])

    penalty = np.eye(design_matrix.shape[1]) * ridge_alpha
    penalty[0, 0] = 0.0
    coefficients = np.linalg.solve(design_matrix.T @ design_matrix + penalty, design_matrix.T @ y_vector)

    home_field_advantage = float(coefficients[0])
    feature_coefficients = coefficients[1:]

    predicted_margins = design_matrix @ coefficients
    errors = predicted_margins - y_vector
    mae = float(np.mean(np.abs(errors)))
    rmse = float(np.sqrt(np.mean(errors**2)))
    correlation = float(np.corrcoef(predicted_margins, y_vector)[0, 1])

    diagnostics = {
        "samples": len(samples),
        "home_field_advantage": round(home_field_advantage, 3),
        "mae": round(mae, 3),
        "rmse": round(rmse, 3),
        "correlation": round(correlation, 3),
    }

    for index, sample in enumerate(samples):
        model_home_margin = float(predicted_margins[index])
        model_home_line = -model_home_margin
        sample["model_home_margin"] = round(model_home_margin, 3)
        sample["model_home_line"] = round(model_home_line, 3)
        sample["error_points"] = round(model_home_line - sample["market_home_line"], 3)
        sample["abs_error_points"] = round(abs(sample["error_points"]), 3)

    return feature_coefficients, home_field_advantage, samples, team_vectors, diagnostics


def component_score(
    team_vectors: dict[str, np.ndarray],
    coeffs_by_feature: dict[str, float],
    subset: list[str],
    calibration_features: list[str] = DEFAULT_CALIBRATION_FEATURES,
) -> dict[str, float]:
    subset_indices = [calibration_features.index(feature) for feature in subset]
    values = {
        team: float(sum(vector[index] * coeffs_by_feature[calibration_features[index]] for index in subset_indices))
        for team, vector in team_vectors.items()
    }
    raw = list(values.values())
    minimum = min(raw)
    maximum = max(raw)
    if math.isclose(minimum, maximum):
        return {team: 50.0 for team in values}
    return {team: ((value - minimum) / (maximum - minimum)) * 100 for team, value in values.items()}


def blended_team_ratings(
    rows: list[dict[str, Any]],
    feature_coefficients: np.ndarray,
    team_vectors: dict[str, np.ndarray],
    calibration_features: list[str] = DEFAULT_CALIBRATION_FEATURES,
    prior_decay_games: float = PRIOR_DECAY_GAMES,
    min_prior_weight: float = MIN_PRIOR_WEIGHT,
    prior_scale: float = PRIOR_SCALE,
) -> dict[str, float]:
    coeffs_by_feature = {
        feature: float(feature_coefficients[index])
        for index, feature in enumerate(calibration_features)
    }
    raw_ratings = {team: float(vector @ feature_coefficients) for team, vector in team_vectors.items()}
    raw_mean = sum(raw_ratings.values()) / len(raw_ratings)
    raw_ratings = {team: rating - raw_mean for team, rating in raw_ratings.items()}

    prior_source = component_score(
        team_vectors,
        coeffs_by_feature,
        ["avg_team_pregame_elo", "talent"],
        calibration_features=calibration_features,
    )
    prior_scaled = {team: (score - 50.0) / 50.0 * prior_scale for team, score in prior_source.items()}

    blended: dict[str, float] = {}
    for row in rows:
        team = str(row["team"])
        fbs_games = float(row.get("fbs_games", 0) or 0)
        prior_weight = max(min_prior_weight, math.exp(-fbs_games / prior_decay_games))
        blended[team] = (prior_weight * prior_scaled[team]) + ((1 - prior_weight) * raw_ratings[team])

    blended_mean = sum(blended.values()) / len(blended)
    return {team: rating - blended_mean for team, rating in blended.items()}


def build_rankings(
    rows: list[dict[str, Any]],
    feature_coefficients: np.ndarray,
    team_vectors: dict[str, np.ndarray],
    calibration_features: list[str] = DEFAULT_CALIBRATION_FEATURES,
) -> list[dict[str, Any]]:
    coeffs_by_feature = {
        feature: float(feature_coefficients[index])
        for index, feature in enumerate(calibration_features)
    }
    ratings = blended_team_ratings(rows, feature_coefficients, team_vectors, calibration_features=calibration_features)

    efficiency_scores = component_score(
        team_vectors,
        coeffs_by_feature,
        [
            "offense_ppa",
            "defense_ppa_allowed",
            "offense_success_rate",
            "defense_success_rate_allowed",
            "wepa_total",
            "wepa_allowed_total",
        ],
        calibration_features=calibration_features,
    )
    market_scores = component_score(
        team_vectors,
        coeffs_by_feature,
        ["avg_team_postgame_elo", "avg_team_pregame_elo", "avg_cover_margin", "talent"],
        calibration_features=calibration_features,
    )
    schedule_scores = component_score(
        team_vectors,
        coeffs_by_feature,
        ["fbs_avg_margin", "avg_opponent_postgame_elo", "fbs_avg_opponent_pregame_elo"],
        calibration_features=calibration_features,
    )

    rankings: list[dict[str, Any]] = []
    for row in rows:
        team = str(row["team"])
        rankings.append(
            {
                "team": team,
                "conference": row.get("conference", ""),
                "record": f"{row.get('wins', 0)}-{row.get('losses', 0)}",
                "fbs_record": f"{row.get('fbs_wins', 0)}-{row.get('fbs_losses', 0)}",
                "rating": round(float(ratings[team]), 2),
                "efficiency_score": round(efficiency_scores[team], 2),
                "market_score": round(market_scores[team], 2),
                "schedule_score": round(schedule_scores[team], 2),
                "avg_team_postgame_elo": round(float(row.get("avg_team_postgame_elo", 0.0)), 2),
                "fbs_win_pct": round(float(row.get("fbs_win_pct", 0.0)), 4),
                "fbs_avg_margin": round(float(row.get("fbs_avg_margin", 0.0)), 3),
                "fbs_avg_opponent_pregame_elo": round(float(row.get("fbs_avg_opponent_pregame_elo", 0.0)), 3),
                "avg_cover_margin": round(float(row.get("avg_cover_margin", 0.0)), 3),
                "offense_ppa": round(float(row.get("offense_ppa", 0.0)), 6),
                "defense_ppa_allowed": round(float(row.get("defense_ppa_allowed", 0.0)), 6),
                "wepa_total": round(float(row.get("wepa_total", 0.0)), 6),
                "wepa_allowed_total": round(float(row.get("wepa_allowed_total", 0.0)), 6),
                "talent": round(float(row.get("talent", 0.0)), 3) if row.get("talent") is not None else "",
                "ap_latest_rank": row.get("ap_latest_rank", ""),
                "cfp_latest_rank": row.get("cfp_latest_rank", ""),
            }
        )

    return sorted(rankings, key=lambda row: (-float(row["rating"]), -float(row["market_score"]), str(row["team"])))


def print_rankings(rankings: list[dict[str, Any]], top_n: int) -> None:
    print("College Football Power Ratings")
    print()
    print(f"{'RK':<4}{'TEAM':<18}{'REC':<7}{'NTR':<8}{'EFF':<8}{'MKT':<8}{'SCH':<8}{'ELO':<9}")
    for index, team in enumerate(rankings[:top_n], start=1):
        print(
            f"{index:<4}"
            f"{str(team['team']):<18}"
            f"{str(team['record']):<7}"
            f"{float(team['rating']):<8.2f}"
            f"{float(team['efficiency_score']):<8.2f}"
            f"{float(team['market_score']):<8.2f}"
            f"{float(team['schedule_score']):<8.2f}"
            f"{float(team['avg_team_postgame_elo']):<9.2f}"
        )


def find_team(rankings: list[dict[str, Any]], team_name: str) -> dict[str, Any]:
    normalized_target = team_name.strip().lower()
    for row in rankings:
        if str(row["team"]).lower() == normalized_target:
            return row

    partial_matches = [row for row in rankings if normalized_target in str(row["team"]).lower()]
    if len(partial_matches) == 1:
        return partial_matches[0]
    raise SystemExit(f"Could not uniquely match team name: {team_name}")


def print_matchup(rankings: list[dict[str, Any]], team_a_name: str, team_b_name: str) -> None:
    team_a = find_team(rankings, team_a_name)
    team_b = find_team(rankings, team_b_name)
    spread = float(team_a["rating"]) - float(team_b["rating"])

    favorite = team_a if spread >= 0 else team_b
    underdog = team_b if spread >= 0 else team_a
    favorite_spread = abs(spread)

    print()
    print("Neutral Field Matchup")
    print(f"{favorite['team']} would be favored over {underdog['team']} by {favorite_spread:.1f} points.")
    print(
        f"Ratings: {team_a['team']} {float(team_a['rating']):.2f}, "
        f"{team_b['team']} {float(team_b['rating']):.2f}"
    )


def save_rankings(rankings: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rankings[0].keys()) if rankings else []
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rankings)


def autosize_worksheet(worksheet) -> None:
    for column in worksheet.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column]
        worksheet.column_dimensions[column[0].column_letter].width = min(max(len(value) for value in values) + 2, 28)


def write_excel(
    rankings: list[dict[str, Any]],
    coeffs_by_feature: dict[str, float],
    diagnostics: dict[str, Any],
    samples: list[dict[str, Any]],
    output_path: Path,
) -> None:
    workbook = Workbook()

    rankings_sheet = workbook.active
    rankings_sheet.title = "Rankings"
    rankings_headers = list(rankings[0].keys()) if rankings else []
    rankings_sheet.append(rankings_headers)
    for cell in rankings_sheet[1]:
        cell.font = Font(bold=True)
    for row in rankings:
        rankings_sheet.append([row.get(header) for header in rankings_headers])
    autosize_worksheet(rankings_sheet)

    teams_sheet = workbook.create_sheet("Teams")
    teams_sheet.append(["team", "rating"])
    teams_sheet["A1"].font = Font(bold=True)
    teams_sheet["B1"].font = Font(bold=True)
    for row in rankings:
        teams_sheet.append([row["team"], row["rating"]])
    autosize_worksheet(teams_sheet)

    model_sheet = workbook.create_sheet("Model")
    model_sheet.append(["metric", "value"])
    model_sheet["A1"].font = Font(bold=True)
    model_sheet["B1"].font = Font(bold=True)
    for key, value in diagnostics.items():
        model_sheet.append([key, value])
    model_sheet.append([])
    model_sheet.append(["feature", "coefficient_points"])
    model_sheet[f"A{model_sheet.max_row}"].font = Font(bold=True)
    model_sheet[f"B{model_sheet.max_row}"].font = Font(bold=True)
    for feature, coefficient in sorted(coeffs_by_feature.items(), key=lambda item: abs(item[1]), reverse=True):
        model_sheet.append([feature, round(coefficient, 4)])
    autosize_worksheet(model_sheet)

    calibration_sheet = workbook.create_sheet("Calibration")
    calibration_headers = list(samples[0].keys()) if samples else []
    calibration_sheet.append(calibration_headers)
    for cell in calibration_sheet[1]:
        cell.font = Font(bold=True)
    for sample in sorted(samples, key=lambda row: row["abs_error_points"], reverse=True):
        calibration_sheet.append([sample.get(header) for header in calibration_headers])
    autosize_worksheet(calibration_sheet)

    matchup_sheet = workbook.create_sheet("Matchup")
    matchup_sheet["A1"] = "Neutral-Field Matchup Tool"
    matchup_sheet["A1"].font = Font(bold=True)
    matchup_sheet["A3"] = "Team A"
    matchup_sheet["A4"] = "Team B"
    matchup_sheet["A6"] = "Team A Rating"
    matchup_sheet["A7"] = "Team B Rating"
    matchup_sheet["A9"] = "Projected Spread"
    matchup_sheet["A10"] = "Team A Win Probability"
    matchup_sheet["A11"] = "Team B Win Probability"
    matchup_sheet["A13"] = "How to use"
    matchup_sheet["A13"].font = Font(bold=True)
    matchup_sheet["A14"] = "Pick two teams from the dropdowns. Spread is neutral-field Team A minus Team B."
    matchup_sheet["A15"] = "Win probability assumes game margin is normally distributed around the model spread."

    matchup_sheet["B3"] = rankings[0]["team"] if rankings else ""
    matchup_sheet["B4"] = rankings[1]["team"] if len(rankings) > 1 else ""
    matchup_sheet["B6"] = '=IFERROR(INDEX(Teams!$B:$B,MATCH(B3,Teams!$A:$A,0)),"")'
    matchup_sheet["B7"] = '=IFERROR(INDEX(Teams!$B:$B,MATCH(B4,Teams!$A:$A,0)),"")'
    matchup_sheet["B9"] = '=IF(OR(B6="",B7=""),"",B6-B7)'
    matchup_sheet["B10"] = '=IF(B9="","",NORM.S.DIST(B9/16,TRUE))'
    matchup_sheet["B11"] = '=IF(B10="","",1-B10)'
    matchup_sheet["C9"] = 'Team A favored if positive'

    team_count = len(rankings) + 1
    validation = DataValidation(type="list", formula1=f"=Teams!$A$2:$A${team_count}", allow_blank=False)
    matchup_sheet.add_data_validation(validation)
    validation.add(matchup_sheet["B3"])
    validation.add(matchup_sheet["B4"])
    autosize_worksheet(matchup_sheet)

    teams_sheet.sheet_state = "hidden"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    rows = load_team_features(args.features)
    if not rows:
        raise SystemExit("No team features found in the provided file.")

    lines = load_lines(args.lines)
    feature_coefficients, home_field_advantage, samples, team_vectors, diagnostics = fit_market_model(rows, lines)
    rankings = build_rankings(rows, feature_coefficients, team_vectors)
    print_rankings(rankings, args.top)
    print()
    print(
        "Model fit: "
        f"home field {home_field_advantage:.2f}, "
        f"MAE {diagnostics['mae']:.2f}, "
        f"RMSE {diagnostics['rmse']:.2f}, "
        f"corr {diagnostics['correlation']:.3f}"
    )
    if args.team_a and args.team_b:
        print_matchup(rankings, args.team_a, args.team_b)
    if args.save:
        save_rankings(rankings, args.save)
        print(f"Saved full rankings to {args.save}")
    if args.excel:
        coeffs_by_feature = {
            feature: float(feature_coefficients[index])
            for index, feature in enumerate(DEFAULT_CALIBRATION_FEATURES)
        }
        write_excel(rankings, coeffs_by_feature, diagnostics, samples, args.excel)
        print(f"Saved Excel workbook to {args.excel}")


if __name__ == "__main__":
    main()
